"""Intelligent camera registry import API.

Validation is a dry-run first. Errors block import. Warnings are non-fatal but
require explicit acknowledgement. A fully clean file is green/ready and can
be imported directly. The endpoint returns field/row locations for every issue.
"""
from __future__ import annotations

import csv, io, json, os, uuid
from typing import Any
from openpyxl import load_workbook
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.exc import IntegrityError
from pydantic import ValidationError
from models import Camera, CameraCreate
from auth import require_permission, Principal
from database import get_db
from routes.cameras import _audit_state, _validate_coordinates, _validate_vendor_model
from services.camera_import_intelligence import normalize_headers, analyze_row, summarize

router = APIRouter(prefix="/camera-imports", tags=["camera-imports"])
MAX_IMPORT_BYTES = 5 * 1024 * 1024


def _parse(raw: bytes, suffix: str) -> tuple[list[dict], list[str]]:
    if suffix == ".csv":
        reader = csv.DictReader(io.StringIO(raw.decode("utf-8-sig")))
        return list(reader), reader.fieldnames or []
    sheet = load_workbook(io.BytesIO(raw), read_only=True, data_only=True).active
    values = sheet.iter_rows(values_only=True)
    headers = next(values, None)
    if not headers:
        return [], []
    names = [str(v).strip() if v is not None else "" for v in headers]
    rows = [{names[i]: value for i, value in enumerate(row) if i < len(names) and names[i]}
            for row in values if any(value is not None and str(value).strip() for value in row)]
    return rows, names


async def _analyze(file: UploadFile) -> dict:
    suffix = os.path.splitext(file.filename or "")[1].lower()
    if suffix not in {".csv", ".xlsx"}:
        raise HTTPException(415, "Upload a CSV or XLSX registry file")
    raw = await file.read()
    if len(raw) > MAX_IMPORT_BYTES:
        raise HTTPException(413, "Registry import exceeds the 5 MiB limit")
    try:
        rows, headers = _parse(raw, suffix)
    except (UnicodeDecodeError, ValueError, OSError, KeyError, TypeError, IndexError) as exc:
        raise HTTPException(422, "Registry file could not be parsed or validated; upload a valid CSV/XLSX registry file.") from exc
    if not rows:
        raise HTTPException(422, "Registry file contains no data rows")
    mapping, header_issues = normalize_headers(headers)
    analyses = [analyze_row(row, number, mapping) for number, row in enumerate(rows, start=2)]
    summary = summarize(analyses, header_issues)
    return {"filename": file.filename, "headers": headers, "header_mapping": mapping, "summary": summary, "rows": analyses, "raw": raw, "suffix": suffix}


def _public_analysis(data: dict) -> dict:
    return {k: v for k, v in data.items() if k not in {"raw", "suffix"}}


@router.post("/validate")
async def validate_camera_import(file: UploadFile = File(...), _: Principal = Depends(require_permission("registry:admin"))):
    data = await _analyze(file)
    return _public_analysis(data)


@router.post("/import", status_code=201)
async def import_camera_registry(
    file: UploadFile = File(...),
    acknowledge_warnings: bool = Query(False),
    principal: Principal = Depends(require_permission("registry:admin")),
    db: AsyncSession = Depends(get_db),
):
    data = await _analyze(file)
    summary = data["summary"]
    if not summary["allow_upload"]:
        raise HTTPException(422, detail={"message": "Import blocked by validation errors. Correct the highlighted fields and upload again.", "analysis": _public_analysis(data)})
    if summary["requires_warning_ack"] and not acknowledge_warnings:
        raise HTTPException(409, detail={"message": "This registry contains warnings. Review them and explicitly acknowledge before importing.", "analysis": _public_analysis(data)})

    import_id = uuid.uuid4()
    await db.execute(text("INSERT INTO camera_imports(id, filename, actor, total_rows) VALUES (CAST(:id AS uuid), :filename, :actor, :total)"), {"id": str(import_id), "filename": data["filename"] or "camera-import.csv", "actor": principal.username, "total": summary["total_rows"]})
    accepted, errors = 0, []
    for analysis_row in data["rows"]:
        if analysis_row["status"] == "blocked":
            errors.extend(analysis_row["issues"]); continue
        try:
            payload = dict(analysis_row["normalized"])
            payload.setdefault("source_system", "csv")
            if "lat" in payload and "lng" in payload:
                payload.setdefault("coord_source", "csv"); payload.setdefault("coord_confidence", 0.7)
            body = CameraCreate.model_validate(payload)
            _validate_coordinates(body)
            await _validate_vendor_model(body, db)
            existing = None
            if body.stream_id is not None:
                existing = await db.scalar(select(Camera).where(Camera.stream_id == body.stream_id))
            elif body.external_id:
                existing = await db.scalar(select(Camera).where(Camera.source_system == body.source_system, Camera.external_id == body.external_id))
            elif body.rtsp_url:
                existing = await db.scalar(select(Camera).where(Camera.rtsp_url == body.rtsp_url))
            values = body.model_dump(exclude_none=True, exclude_unset=True)
            before = _audit_state(existing)
            async with db.begin_nested():
                if existing:
                    if existing.coord_source == "manual" and (existing.coord_confidence or 0) >= 0.9:
                        for key in ("lat", "lng", "coord_source", "coord_confidence"): values.pop(key, None)
                    for key, value in values.items(): setattr(existing, key, value)
                    camera, action = existing, "bulk_update"
                else:
                    camera, action = Camera(**values), "bulk_create"; db.add(camera)
                await db.flush(); await db.refresh(camera)
                await db.execute(text("""INSERT INTO camera_audit_log(camera_id, actor, action, before_value, after_value, correlation_id)
                    VALUES (CAST(:id AS uuid), :actor, :action, CAST(:before AS jsonb), CAST(:after AS jsonb), CAST(:correlation AS uuid))"""), {
                    "id": str(camera.id), "actor": principal.username, "action": action, "before": json.dumps(before),
                    "after": json.dumps({**_audit_state(camera), "import_id": str(import_id), "quality": analysis_row["status"]}), "correlation": str(import_id)})
            accepted += 1
        except (ValidationError, ValueError, TypeError, IntegrityError) as exc:
            errors.append({"row": analysis_row["row"], "severity": "error", "field": "row", "message": str(exc)})
    rejected = summary["total_rows"] - accepted
    await db.execute(text("""UPDATE camera_imports SET accepted_rows=:accepted, rejected_rows=:rejected, errors=CAST(:errors AS jsonb), column_map=CAST(:column_map AS jsonb), status='completed', completed_at=NOW() WHERE id=CAST(:id AS uuid)"""), {"id": str(import_id), "accepted": accepted, "rejected": rejected, "errors": json.dumps(errors[:100]), "column_map": json.dumps(data["header_mapping"])})
    await db.commit()
    return {"import_id": str(import_id), "total_rows": summary["total_rows"], "accepted_rows": accepted, "rejected_rows": rejected, "errors": errors[:100], "quality": summary}
