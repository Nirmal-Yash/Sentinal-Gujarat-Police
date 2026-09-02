from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import Response
from sqlalchemy import select, text, or_, String
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from models import Camera, CameraOut, CameraCreate
from auth import require_authenticated, require_role, Principal
from database import get_db
import uuid, os, base64, csv, io, json
from openpyxl import load_workbook
from pydantic import ValidationError
import redis as redis_lib

router = APIRouter(prefix="/cameras", tags=["cameras"], dependencies=[Depends(require_authenticated)])
REDIS_URL = os.getenv("REDIS_URL", "redis://localhost:6379")


def _validate_coordinates(body: CameraCreate):
    if (body.lat is None) != (body.lng is None):
        raise HTTPException(422, "latitude and longitude must be supplied together")


def _audit_state(camera: Camera | None):
    if camera is None:
        return None
    value = CameraOut.model_validate(camera).model_dump(mode="json")
    for key in ("rtsp_url", "hls_url", "whep_url", "stream_url"):
        value.pop(key, None)
    return value


async def _validate_vendor_model(body: CameraCreate, db: AsyncSession):
    """Prevent a camera from referencing a model owned by another vendor."""
    if body.model_id and not body.vendor_id:
        raise HTTPException(422, "vendor_id is required when model_id is supplied")
    if body.vendor_id and not await db.scalar(text("SELECT 1 FROM vendors WHERE id=CAST(:id AS uuid)"), {"id": str(body.vendor_id)}):
        raise HTTPException(422, "vendor_id is not registered")
    if body.model_id:
        owner = await db.scalar(text("SELECT vendor_id FROM camera_models WHERE id=CAST(:id AS uuid)"), {"id": str(body.model_id)})
        if owner is None or str(owner) != str(body.vendor_id):
            raise HTTPException(422, "model_id does not belong to vendor_id")


CSV_ALIASES = {
    "camera_id": "external_id", "id": "external_id", "camera_name": "name", "camera": "name",
    "latitude": "lat", "longitude": "lng", "lon": "lng", "owner": "owner_organization",
    "ownership": "owner_organization", "rtsp": "rtsp_url", "hls": "hls_url", "source": "source_system",
}

def _coordinate(value: str) -> float:
    """Accept decimal coordinates and common DMS notation from field surveys."""
    raw = value.strip().upper().replace("°", " ").replace("'", " ").replace('"', " ")
    direction = -1 if raw.endswith(("S", "W")) else 1
    raw = raw.rstrip("NSEW ")
    parts = [part for part in raw.replace(",", ".").split() if part]
    numbers = [float(part) for part in parts]
    if not numbers: raise ValueError("empty coordinate")
    result = numbers[0] + (numbers[1] / 60 if len(numbers) > 1 else 0) + (numbers[2] / 3600 if len(numbers) > 2 else 0)
    return result * direction

def _csv_payload(row: dict) -> tuple[dict, dict]:
    """Normalize real-world headers; preserve an explicit mapping for audit."""
    payload, column_map = {}, {}
    for key, value in row.items():
        normalized = (key or "").strip().lower().replace(" ", "_").replace("-", "_")
        target = CSV_ALIASES.get(normalized, normalized)
        if value is not None and value.strip() != "":
            payload[target] = value.strip(); column_map[key] = target
    for key in ("stream_id", "retention_days"):
        if key in payload:
            payload[key] = int(payload[key])
    for key in ("lat", "lng"):
        if key in payload:
            payload[key] = _coordinate(payload[key])
    if "analytics_capabilities" in payload:
        payload["analytics_capabilities"] = [item.strip() for item in payload["analytics_capabilities"].split("|") if item.strip()]
    return payload, column_map


@router.get("/", response_model=list[CameraOut])
async def list_cameras(
    q: str | None = Query(None, min_length=1, max_length=100),
    department: str | None = None, status: str | None = None,
    health_status: str | None = None, camera_type: str | None = None,
    vendor_id: uuid.UUID | None = None, model_id: uuid.UUID | None = None,
    limit: int = Query(250, ge=1, le=500), offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
):
    """Registry read model.  It is the only camera metadata source for UI/GIS."""
    stmt = select(Camera).where(Camera.status != 'deleted')
    if q:
        term = f"%{q.strip()}%"
        stmt = stmt.where(or_(Camera.name.ilike(term), Camera.location.ilike(term),
                              Camera.department.ilike(term), Camera.owner_organization.ilike(term),
                              Camera.status.ilike(term), Camera.health_status.ilike(term),
                              Camera.camera_type.ilike(term),
                              Camera.stream_id.cast(String).ilike(term)))
    for column, value in ((Camera.department, department), (Camera.status, status),
                          (Camera.health_status, health_status), (Camera.camera_type, camera_type)):
        if value:
            stmt = stmt.where(column.ilike(value))
    if vendor_id: stmt = stmt.where(Camera.vendor_id == vendor_id)
    if model_id: stmt = stmt.where(Camera.model_id == model_id)
    result = await db.execute(stmt.order_by(Camera.stream_id).limit(limit).offset(offset))
    return result.scalars().all()


@router.post("/", response_model=CameraOut, status_code=201)
@router.post("/onboard", response_model=CameraOut, status_code=201)
async def onboard_camera(body: CameraCreate, principal: Principal = Depends(require_role("ADMIN")), db: AsyncSession = Depends(get_db)):
    """Manual/API Model-1 onboarding; catalogue sync remains the same owner for external sources."""
    _validate_coordinates(body)
    await _validate_vendor_model(body, db)
    if body.stream_id is not None and await db.scalar(select(Camera.id).where(Camera.stream_id == body.stream_id)):
        raise HTTPException(409, "stream_id is already registered")
    if body.rtsp_url and await db.scalar(select(Camera.id).where(Camera.rtsp_url == body.rtsp_url)):
        raise HTTPException(409, "RTSP source is already registered")
    camera = Camera(**body.model_dump())
    db.add(camera)
    await db.flush()
    if body.lat is not None:
        await db.execute(text("UPDATE cameras SET geom=ST_SetSRID(ST_MakePoint(:lng, :lat), 4326), updated_at=NOW() WHERE id=:id"),
                         {"id": str(camera.id), "lat": body.lat, "lng": body.lng})
    await db.refresh(camera)
    await db.execute(text("""INSERT INTO camera_audit_log(camera_id, actor, action, before_value, after_value, correlation_id)
        VALUES (CAST(:id AS uuid), :actor, 'create', NULL, CAST(:value AS jsonb), CAST(:correlation AS uuid))"""),
        {"id": str(camera.id), "actor": principal.username, "value": json.dumps(_audit_state(camera)), "correlation": str(uuid.uuid4())})
    await db.commit()
    await db.refresh(camera)
    return camera


@router.post("/imports/csv", status_code=201)
async def import_cameras_csv(
    file: UploadFile = File(...), principal: Principal = Depends(require_role("ADMIN")),
    db: AsyncSession = Depends(get_db),
):
    """Controlled Model-1 bulk onboarding; invalid rows are reported, never dropped."""
    suffix = os.path.splitext(file.filename or "")[1].lower()
    if suffix not in {".csv", ".xlsx"}:
        raise HTTPException(415, "Upload a CSV or XLSX registry file")
    raw = await file.read()
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(413, "Registry import exceeds the 5 MiB PoC limit")
    try:
        if suffix == ".csv":
            rows = list(csv.DictReader(io.StringIO(raw.decode("utf-8-sig"))))
        else:
            sheet = load_workbook(io.BytesIO(raw), read_only=True, data_only=True).active
            values = sheet.iter_rows(values_only=True)
            headers = next(values, None)
            if not headers:
                rows = []
            else:
                header_names = [str(value).strip() if value is not None else "" for value in headers]
                rows = [{header_names[index]: value for index, value in enumerate(row) if index < len(header_names) and header_names[index]}
                        for row in values if any(value is not None and str(value).strip() for value in row)]
    except (UnicodeDecodeError, ValueError, OSError) as exc:
        raise HTTPException(422, "Registry file could not be parsed; CSV must be UTF-8 and XLSX must have a header row") from exc
    if not rows:
        raise HTTPException(422, "CSV contains no data rows")

    import_id = uuid.uuid4()
    await db.execute(text("""INSERT INTO camera_imports(id, filename, actor, total_rows)
        VALUES (CAST(:id AS uuid), :filename, :actor, :total)"""),
        {"id": str(import_id), "filename": file.filename or "camera-import.csv", "actor": principal.username, "total": len(rows)})
    accepted, errors, import_column_map = 0, [], {}
    for row_number, row in enumerate(rows, start=2):
        try:
            payload, mapping = _csv_payload(row)
            import_column_map.update(mapping)
            source_supplied = "source_system" in payload
            payload.setdefault("source_system", "csv")
            if "lat" in payload and "lng" in payload:
                payload.setdefault("coord_source", "csv")
                payload.setdefault("coord_confidence", 0.7)
            body = CameraCreate.model_validate(payload)
            _validate_coordinates(body)
            await _validate_vendor_model(body, db)
            # Government files are often incomplete.  Prefer a stable source
            # key, then an RTSP URL; a row with neither is still imported as a
            # clearly auditable new asset rather than silently discarded.
            existing = None
            if body.stream_id is not None:
                existing = await db.scalar(select(Camera).where(Camera.stream_id == body.stream_id))
            elif body.external_id:
                existing = await db.scalar(select(Camera).where(Camera.source_system == body.source_system, Camera.external_id == body.external_id))
            elif body.rtsp_url:
                existing = await db.scalar(select(Camera).where(Camera.rtsp_url == body.rtsp_url))
            # Sparse government files must not overwrite registry fields with
            # Pydantic defaults such as an empty location.
            values = body.model_dump(exclude_none=True, exclude_unset=True)
            if existing and not source_supplied:
                values.pop("source_system", None)
            before_value = _audit_state(existing)
            async with db.begin_nested():
                if existing:
                    # Imported coordinates never replace a manually verified point.
                    if existing.coord_source == "manual" and (existing.coord_confidence or 0) >= 0.9:
                        values.pop("lat", None); values.pop("lng", None); values.pop("coord_source", None); values.pop("coord_confidence", None)
                    for key, value in values.items():
                        setattr(existing, key, value)
                    camera, action = existing, "bulk_update"
                else:
                    camera, action = Camera(**values), "bulk_create"
                    db.add(camera)
                await db.flush()
                await db.refresh(camera)
                await db.execute(text("""INSERT INTO camera_audit_log(camera_id, actor, action, before_value, after_value, correlation_id)
                    VALUES (CAST(:id AS uuid), :actor, :action, CAST(:before AS jsonb), CAST(:after AS jsonb), CAST(:correlation AS uuid))"""),
                    {"id": str(camera.id), "actor": principal.username, "action": action,
                     "before": json.dumps(before_value), "after": json.dumps({**_audit_state(camera), "import_id": str(import_id), "column_map": mapping}),
                     "correlation": str(import_id)})
            accepted += 1
        except (ValidationError, ValueError, TypeError, IntegrityError) as exc:
            errors.append({"row": row_number, "error": str(exc)})
    await db.execute(text("""UPDATE camera_imports SET accepted_rows=:accepted, rejected_rows=:rejected,
        errors=CAST(:errors AS jsonb), column_map=CAST(:column_map AS jsonb), status='completed', completed_at=NOW() WHERE id=CAST(:id AS uuid)"""),
        {"id": str(import_id), "accepted": accepted, "rejected": len(rows) - accepted, "errors": json.dumps(errors[:100]), "column_map": json.dumps(import_column_map)})
    await db.commit()
    return {"import_id": str(import_id), "total_rows": len(rows), "accepted_rows": accepted,
            "rejected_rows": len(rows) - accepted, "errors": errors[:100]}


@router.get("/imports")
async def list_camera_imports(limit: int = Query(20, ge=1, le=100), _: Principal = Depends(require_role("ADMIN")), db: AsyncSession = Depends(get_db)):
    result = await db.execute(text("SELECT * FROM camera_imports ORDER BY created_at DESC LIMIT :limit"), {"limit": limit})
    return [dict(row) for row in result.mappings()]


@router.get("/export")
async def export_cameras(profile: str = Query("registry", pattern="^(registry|health|audit)$"), db: AsyncSession = Depends(get_db)):
    """Export public registry metadata only; URLs/credentials are never exported."""
    result = await db.execute(select(Camera).where(Camera.status != 'deleted').order_by(Camera.stream_id))
    if profile == "audit":
        result = await db.execute(text("SELECT camera_id,actor,action,before_value,after_value,correlation_id,created_at FROM camera_audit_log ORDER BY created_at DESC"))
        headers, values_rows = "camera_id,actor,action,before_value,after_value,correlation_id,created_at\n", result.mappings().all()
        rows = [headers] + [",".join('"' + str(value or '').replace('"', '""') + '"' for value in row.values()) + "\n" for row in values_rows]
        return Response("".join(rows), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=camera-audit.csv"})
    headers = ("id,stream_id,name,location,latitude,longitude,coord_source,coord_confidence,department,owner,camera_type,status,health_status,maintenance_status,retention_days,analytics_capabilities,vendor_id,model_id\n"
               if profile == "registry" else "id,stream_id,name,status,health_status,connectivity_status,last_frame_at,observed_at,source_fps,decode_fps,published_fps,reconnect_count,decode_failure_count\n")
    rows = [headers]
    for c in result.scalars():
        values = ([c.id, c.stream_id, c.name, c.location, c.lat, c.lng, c.coord_source, c.coord_confidence, c.department,
                   c.owner_organization, c.camera_type, c.status, c.health_status, c.maintenance_status, c.retention_days,
                   c.analytics_capabilities, c.vendor_id, c.model_id] if profile == "registry" else
                  [c.id, c.stream_id, c.name, c.status, c.health_status, c.connectivity_status, c.last_frame_at, c.observed_at,
                   c.observed_source_fps, c.observed_decode_fps, c.observed_published_fps, c.reconnect_count, c.decode_failure_count])
        rows.append(",".join('"' + str(v or '').replace('"', '""') + '"' for v in values) + "\n")
    return Response("".join(rows), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=camera-registry.csv"})


@router.get("/geojson")
async def cameras_geojson(db: AsyncSession = Depends(get_db)):
    result = await db.execute(text("""SELECT id, stream_id, name, department, camera_type, status, health_status,
        coord_source, coord_confidence, ST_AsGeoJSON(geom)::json AS geometry
        FROM cameras WHERE status <> 'deleted' AND geom IS NOT NULL ORDER BY stream_id"""))
    return {"type": "FeatureCollection", "features": [
        {"type": "Feature", "id": str(c["id"]), "geometry": json.loads(c["geometry"]) if isinstance(c["geometry"], str) else c["geometry"],
         "properties": {key: c[key] for key in ("id", "name", "stream_id", "department", "camera_type", "status", "health_status", "coord_source", "coord_confidence")}}
        for c in result.mappings()
    ]}


@router.get("/stats/summary")
async def camera_stats(db: AsyncSession = Depends(get_db)):
    result = await db.execute(text("""
        SELECT c.id, c.name, c.stream_id, c.status, c.codec,
               c.width, c.height, c.hls_url,
               COUNT(a.id) FILTER (WHERE a.created_at > NOW() - INTERVAL '1 hour') AS alerts_1h,
               COUNT(a.id) FILTER (WHERE a.acknowledged = FALSE)                   AS unacked
        FROM cameras c
        LEFT JOIN alerts a ON a.cam_id = c.id
        WHERE c.status != 'deleted'
        GROUP BY c.id, c.name, c.stream_id, c.status, c.codec, c.width, c.height, c.hls_url
        ORDER BY c.stream_id
    """))
    return [dict(r) for r in result.mappings().all()]


@router.get("/analytics/recent")
async def recent_camera_analytics(
    seconds: int = Query(90, ge=10, le=3600),
    db: AsyncSession = Depends(get_db),
):
    """Latest persisted analytics per camera for live UI overlays.

    This deliberately reads durable detections instead of a transient Redis
    consumer stream, so a rendered plate is auditable and searchable later.
    """
    result = await db.execute(text("""
        SELECT DISTINCT ON (cam_id)
               id, cam_id, detection_type, plate_text, confidence, timestamp,
               global_track_id, track_id
        FROM detections
        WHERE cam_id IS NOT NULL
          AND timestamp >= NOW() - (CAST(:seconds AS integer) * INTERVAL '1 second')
          AND plate_text IS NOT NULL AND plate_text <> ''
        ORDER BY cam_id, timestamp DESC
    """), {"seconds": seconds})
    return [dict(row) for row in result.mappings().all()]


@router.get("/{cam_id}", response_model=CameraOut)
async def get_camera(cam_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    cam = await db.get(Camera, cam_id)
    if not cam:
        raise HTTPException(404, "Camera not found")
    return cam


@router.get("/{cam_id}/snapshot")
async def snapshot(cam_id: uuid.UUID):
    """Return latest JPEG frame cached by ingestion worker (10 s TTL)."""
    r    = redis_lib.from_url(REDIS_URL, decode_responses=False)
    data = r.get(f"snapshot:{cam_id}")
    if not data:
        raise HTTPException(404, "No snapshot available yet — stream may still be connecting")
    return Response(content=base64.b64decode(data), media_type="image/jpeg")


@router.get("/pipeline/stats")
async def pipeline_stats():
    """Return Redis stream lengths — proves pipeline is alive without DB queries."""
    import redis as redis_lib
    try:
        r = redis_lib.from_url(REDIS_URL, decode_responses=True)
        return {
            "raw_frames":  r.xlen("raw_frames"),
            "detections":  r.xlen("detections"),
            "alerts":      r.xlen("alerts"),
            "cam_resets":  r.xlen("cam_resets") if r.exists("cam_resets") else 0,
        }
    except Exception as e:
        return {"error": str(e)}
